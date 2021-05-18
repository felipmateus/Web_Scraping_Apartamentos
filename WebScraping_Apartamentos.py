from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.keys import Keys
from time import sleep
import openpyxl


class scrappy:

    def iniciar(self):
        self.driver: WebDriver = webdriver.Chrome()
        url = "https://www.zapimoveis.com.br/"
        self.driver.get(url)
        self.driver.find_element_by_xpath('//*[@id="app"]/section/section[1]/div/section/form/div/div[1]/div[1]/div/button[2]').click()
        self.driver.find_element_by_css_selector('#l-select1 > optgroup:nth-child(2) > option:nth-child(1)').click()
        self.driver.find_element_by_xpath('//*[@id="app"]/section/section[1]/div/section/form/div/div[2]/div/div/div/input').send_keys('Niterói, Icarai')
        self.driver.find_element_by_xpath('//*[@id="app"]/section/section[1]/div/section/form/div/div[2]/button').send_keys(Keys.ENTER)
        sleep(2)

    def raspagem_dados(self):

        cont = 1
        for i in range(10):
            try:
                self.valores_aluguel_text = []
                valores_aluguel = self.driver.find_elements_by_xpath('//p[@class="simple-card__price js-price heading-regular heading-regular__bolder align-left"]')
                for valor_aluguel in valores_aluguel:
                    self.valores_aluguel_text.append(valor_aluguel.text)
                #  sleep(1)
            except:
                print("Não foi possível raspar os valores de aluguel da página " + str(cont))

            try:
                self.valores_iptu_text = []
                valores_iptu = self.driver.find_elements_by_xpath("//li[@class='card-price__item iptu text-regular']")
                for valor_iptu in valores_iptu:
                    self.valores_iptu_text.append(valor_iptu.text)
                # sleep(1)
            except:
                print("Não foi possível obter os valores de IPTU da página " + str(cont))

            # try:
            #     self.descricoes_text = []
            #     descricoes_apartamentos = self.driver.find_elements_by_xpath('//div[@class="collapse simple-card__description"]')
            #     for descricao_apartamento in descricoes_apartamentos:
            #         self.descricoes_text.append(descricao_apartamento.text)
                # sleep(1)
            #
            # except:
            #     print("Não foi possível obter as descrições dos apartamentos da página " + str(cont))
            #
            try:
                self.valores_condominio_text = []
                valores_condominio = self.driver.find_elements_by_xpath('//li[@class="card-price__item condominium text-regular"]')
                for valor_condominio in valores_condominio:
                    self.valores_condominio_text.append(valor_condominio.text)
                # sleep(1)
            except:
                print("Não foi possível obter os valores dos condomínios da página " + str(cont))

            try:
                self.enderecos_text = []
                enderecos = self.driver.find_elements_by_xpath('//p[@class="color-dark text-regular simple-card__address"]')
                for endereco in enderecos:
                    self.enderecos_text.append(endereco.text)
                # sleep(1)
            except:
                print("Não foi possível obter os endereços da página " + str(cont))

            try:
                self.tamanho_m2_apartamentos_text = []
                tamanho_m2_apartamentos = self.driver.find_elements_by_xpath('//li[@class="feature__item text-small js-areas"]')
                for tamanho_m2_apartamento in tamanho_m2_apartamentos:
                    self.tamanho_m2_apartamentos_text.append(tamanho_m2_apartamento.text)
                # sleep(1)
            except:
                print("Não foi possível obter as áreas dos apartamentos da página " + str(cont))

            try:
                self.numero_quartos_text = []
                numero_quartos = self.driver.find_elements_by_xpath('//li[@class="feature__item text-small js-bedrooms"]')
                for numero_quarto in numero_quartos:
                    self.numero_quartos_text.append(numero_quarto.text)
                # sleep(1)
            except:
                print("Não foi possível obter os números dos quartos da página" + str(cont))

            try:
                self.numero_vagas_text = []
                numero_vagas = self.driver.find_elements_by_xpath('//li[@class="feature__item text-small js-parking-spaces"]')
                for numero_vaga in numero_vagas:
                    self.numero_vagas_text.append(numero_vaga.text)
                sleep(1)
            except:
                print("Não foi possível obter os números das vagas da página" + str(cont))

            try:
                self.numero_banheiros_text = []
                numero_banheiros = self.driver.find_elements_by_xpath('//li[@class="feature__item text-small js-bathrooms"]')
                for numero_banheiro in numero_banheiros:
                    self.numero_banheiros_text.append(numero_banheiro.text)
                # sleep(1)
            except:
                print("Não foi possível obter os números dos baneheiros da página" + str(cont))

            try:
                cont = cont + 1
                self.driver.find_element_by_xpath('//button[@class="pagination__button pagination__button--next js-next-page button button-primary button-primary--outline button--regular button--icon"]').send_keys(Keys.ENTER)
                print("Próxima página")
                sleep(1)
            except:
                print("Não há mais páginas")

    def criar_planilha(self):
        wb = openpyxl.Workbook()

        wb['Sheet'].title = "Report Automation"

        sh1 = wb.active

        index1 = 1
        for i in self.valores_aluguel_text:
            sh1.cell(column=1, row=index1, value=i)
            index1 = index1 + 1

        index2 = 1
        for valor_iptu in self.valores_iptu_text:
            sh1.cell(column=2, row=index2, value=valor_iptu)
            index2 = index2 + 1

        index3 = 1
        for valor_condomio in self.valores_condominio_text:
            sh1.cell(column=3, row=index3, value=valor_condomio)
            index3 = index3 + 1

        index4 = 1
        for endereco in self.enderecos_text:
            sh1.cell(column=4, row=index4, value=endereco)
            index4 = index4 + 1

        index5 = 1
        for tamanho_m2_apartamento in self.tamanho_m2_apartamentos_text:
            sh1.cell(column=5, row=index5, value=tamanho_m2_apartamento)
            index5 = index5 + 1

        index6 = 1
        for numero_quarto in self.numero_quartos_text:
            sh1.cell(column=6, row=index6, value=numero_quarto)
            index6 = index6 + 1

        index7 = 1
        for numero_vaga in self.numero_vagas_text:
            sh1.cell(column=7, row=index7, value=numero_vaga)
            index7 = index7 + 1

        index8 = 1
        for numero_banheiro in self.numero_banheiros_text:
            sh1.cell(column=8, row=index8, value=numero_banheiro)
            index8 = index8 + 1

        wb.save("C:\\Users\\Felipe\\PycharmProjects\\webscraping_apartamento\\treinando_excel2.xlsx")



start = scrappy()
start.iniciar()
start.raspagem_dados()
start.criar_planilha()

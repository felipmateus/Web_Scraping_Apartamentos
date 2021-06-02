import openpyxl


# import WebScraping_Apartamentos as ws


def criar_planilha_excel(valores_aluguel_text, valores_iptu_text, valores_condominio_text, enderecos_text, tamanho_m2_apartamentos_text, numero_quartos_text, numero_vagas_text, numero_banheiros_text):
    wb = openpyxl.Workbook()

    wb['Sheet'].title = "Report Automation"

    sh1 = wb.active

    index1 = 1
    for i in valores_aluguel_text:
        sh1.cell(column=1, row=index1, value=i)
        index1 = index1 + 1

    index2 = 1
    for valor_iptu in valores_iptu_text:
        sh1.cell(column=2, row=index2, value=valor_iptu)
        index2 = index2 + 1

    index3 = 1
    for valor_condomio in valores_condominio_text:
        sh1.cell(column=3, row=index3, value=valor_condomio)
        index3 = index3 + 1

    index4 = 1
    for endereco in enderecos_text:
        sh1.cell(column=4, row=index4, value=endereco)
        index4 = index4 + 1

    index5 = 1
    for tamanho_m2_apartamento in tamanho_m2_apartamentos_text:
        sh1.cell(column=5, row=index5, value=tamanho_m2_apartamento)
        index5 = index5 + 1

    index6 = 1
    for numero_quarto in numero_quartos_text:
        sh1.cell(column=6, row=index6, value=numero_quarto)
        index6 = index6 + 1

    index7 = 1
    for numero_vaga in numero_vagas_text:
        sh1.cell(column=7, row=index7, value=numero_vaga)
        index7 = index7 + 1

    index8 = 1
    for numero_banheiro in numero_banheiros_text:
        sh1.cell(column=8, row=index8, value=numero_banheiro)
        index8 = index8 + 1

    # Lembra de alterar o teu caminho
    wb.save("C:\\Users\\Felipe\\PycharmProjects\\webscraping_apartamento\\teste.xlsx")

    return(valores_aluguel_text, valores_iptu_text, valores_condominio_text, enderecos_text, tamanho_m2_apartamentos_text, numero_quartos_text, numero_vagas_text, numero_banheiros_text)